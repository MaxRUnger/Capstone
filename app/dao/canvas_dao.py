"""
CanvasDAO — helpers for Canvas integration.

Responsibilities:
1. Parse the Mark Clark template CSV/Excel to bulk-create LOs with vendor codes
2. Generate the Canvas-compatible LO template for download
"""

import csv
import io
import openpyxl
from app.authentication import supabase_admin as supabase
from app.dao.learning_objective_dao import LearningObjectiveDAO


class CanvasDAO:

    # Expected template columns (case-insensitive):
    #   vendor_code | name | homework_group | description
    REQUIRED_COLS = {"vendor_code", "name"}

    @staticmethod
    def parse_template_csv(file_stream, class_id: str) -> tuple[list[dict], list[str]]:
        """
        Parse a CSV template file uploaded by the instructor.
        Returns (lo_list, errors).
        lo_list: list of dicts ready for LearningObjectiveDAO.create_bulk()
        errors:  list of human-readable problem strings
        """
        text = file_stream.read().decode("utf-8")
        reader = csv.DictReader(io.StringIO(text))
        # Normalize headers
        headers = {h.strip().lower(): h for h in (reader.fieldnames or [])}
        errors = []
        for required in CanvasDAO.REQUIRED_COLS:
            if required not in headers:
                errors.append(f"Missing required column: '{required}'")
        if errors:
            return [], errors

        lo_list = []
        for i, row in enumerate(reader, start=2):  # row 1 = header
            norm = {k.strip().lower(): v.strip() for k, v in row.items()}
            vendor_code = norm.get("vendor_code", "").strip()
            name = norm.get("name", "").strip()
            if not vendor_code or not name:
                errors.append(f"Row {i}: missing vendor_code or name — skipped")
                continue
            lo_list.append({
                "vendor_code": vendor_code,
                "name": name,
                "homework_group": norm.get("homework_group", ""),
                "description": norm.get("description", ""),
            })

        return lo_list, errors

    @staticmethod
    def parse_template_excel(file_stream, class_id: str) -> tuple[list[dict], list[str]]:
        """
        Same as parse_template_csv but for .xlsx/.xls files.
        """
        wb = openpyxl.load_workbook(file_stream)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return [], ["File is empty"]

        # First row = headers
        raw_headers = [str(h).strip().lower() if h else "" for h in rows[0]]
        errors = []
        for required in CanvasDAO.REQUIRED_COLS:
            if required not in raw_headers:
                errors.append(f"Missing required column: '{required}'")
        if errors:
            return [], errors

        col_idx = {h: i for i, h in enumerate(raw_headers)}
        lo_list = []
        for i, row in enumerate(rows[1:], start=2):
            vendor_code = str(row[col_idx["vendor_code"]] or "").strip()
            name = str(row[col_idx["name"]] or "").strip()
            if not vendor_code or not name:
                errors.append(f"Row {i}: missing vendor_code or name — skipped")
                continue
            hw_group = str(row[col_idx.get("homework_group", -1)] or "").strip() \
                if "homework_group" in col_idx else ""
            description = str(row[col_idx.get("description", -1)] or "").strip() \
                if "description" in col_idx else ""
            lo_list.append({
                "vendor_code": vendor_code,
                "name": name,
                "homework_group": hw_group,
                "description": description,
            })

        return lo_list, errors

    @staticmethod
    def generate_blank_template_csv() -> str:
        """
        Returns a CSV string for the downloadable blank template.
        Instructors fill this in from Mark Clark's email and upload it.
        """
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(["vendor_code", "name", "homework_group", "description"])
        writer.writerow(["D1_MATH130_001", "LO 1: Example objective", "HW1", "Optional description"])
        writer.writerow(["D1_MATH130_002", "LO 2: Example objective", "HW1", ""])
        writer.writerow(["D1_MATH130_003", "LO 3: Example objective", "HW2", ""])
        return output.getvalue()

    @staticmethod
    def generate_blank_template_xlsx() -> bytes:
        """Returns an XLSX byte string for the downloadable blank template."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Learning Objectives"

        # Header row
        headers = ["vendor_code", "name", "homework_group", "description"]
        ws.append(headers)

        # Example rows
        ws.append(["D1_MATH130_001", "LO 1: Example objective", "HW1", "Optional description"])
        ws.append(["D1_MATH130_002", "LO 2: Example objective", "HW1", ""])
        ws.append(["D1_MATH130_003", "LO 3: Example objective", "HW2", ""])

        # Style header row bold
        from openpyxl.styles import Font, PatternFill, Alignment
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="2563EB")  # blue
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # Column widths
        ws.column_dimensions["A"].width = 22
        ws.column_dimensions["B"].width = 45
        ws.column_dimensions["C"].width = 16
        ws.column_dimensions["D"].width = 40

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()
